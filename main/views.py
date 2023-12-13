from django import template
import uuid
from calendar import month
from cmath import nan
from multiprocessing import cpu_count
from queue import Empty
# from tkinter.messagebox import YES
# from tkinter.tix import Tree
from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
from django.contrib.auth import login, authenticate
from .forms import SignUpForm, AddFollowerForm, AddMonthYearForm
from django.shortcuts import render, redirect
from .models import PublisherProfile, User, MonthYear, PublisherReport
from django.utils.crypto import get_random_string
import datetime
from django.contrib.auth.hashers import make_password
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth import get_user_model
from django.db.models import Sum
from django.db.models import Avg
import csv
# pip install xlwt first
import xlwt

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font

import csv
from io import TextIOWrapper
import xlrd
from django.contrib import messages



User = get_user_model()


register = template.Library()


# Create your views here.
def index(response):
    if response.user.is_authenticated == True:
        return HttpResponseRedirect("/logout")

    return render(response, "main/index.html", {})


def signup_view(request):

    if request.user.is_authenticated == True:
        return HttpResponseRedirect("/logout")

    form = SignUpForm(request.POST)
    if form.is_valid():
        query_phone = PublisherProfile.objects.filter(
            phone=form.cleaned_data.get('phone')).count()
        if len(form.cleaned_data.get('first_name')) > 0 and len(form.cleaned_data.get('last_name')) > 0 and len(form.cleaned_data.get('gender')) > 0 and len(form.cleaned_data.get('phone')) == 11 and form.cleaned_data.get('phone').isdigit() and query_phone == 0 and form.cleaned_data.get('phone') == form.cleaned_data.get('phone2') and len(form.cleaned_data.get('password1')) > 0 and form.cleaned_data.get('password1') == form.cleaned_data.get('password2'):
            user = form.save()
            user.refresh_from_db()
            user.publisherprofile.first_name = form.cleaned_data.get(
                'first_name')
            user.publisherprofile.middle_name = form.cleaned_data.get(
                'middle_name')
            user.publisherprofile.last_name = form.cleaned_data.get(
                'last_name')
            user.publisherprofile.username = form.cleaned_data.get(
                'first_name') + form.cleaned_data.get('username')
            user.publisherprofile.gender = form.cleaned_data.get('gender')
            user.publisherprofile.email = form.cleaned_data.get('email')
            user.publisherprofile.phone = form.cleaned_data.get('phone')
            user.publisherprofile.pioneer = 0
            user.publisherprofile.appointment = 0
            user.publisherprofile.parent_id = form.cleaned_data.get(
                'first_name') + form.cleaned_data.get('username')
            user.publisherprofile.status = 'Pending'
            user.publisherprofile.privilege = 0
            user.publisherprofile.groupName = 0
            user.publisherprofile.created_at = datetime.datetime.now()
            user.save()
            username = form.cleaned_data.get('username')
            phone = form.cleaned_data.get('phone')
            password = form.cleaned_data.get('password1')
            user = authenticate(username=phone, password=password)
            login(request, user)
            return HttpResponseRedirect("/sign-up-suc")
        else:
            return HttpResponseRedirect("/sign-up")
    else:
        form = SignUpForm()
    return render(request, 'main/sign-up.html', {'form': form})


def sign_up_suc(response):
    publisher = response.user

    if publisher.is_authenticated == False:
        return HttpResponseRedirect("/logout")

    if publisher.phone == "":
        return HttpResponseRedirect("/logout")

    if publisher.publisherprofile.status == "Active":
        return HttpResponseRedirect("/logout")

    form = AddFollowerForm(response.POST)
    if form.is_valid():
        phone_value = form.cleaned_data.get('phone')
        query_phone = PublisherProfile.objects.filter(phone=phone_value).exclude(
            parent_id=publisher.publisherprofile.username).count()
        if len(form.cleaned_data.get('first_name')) > 0 and len(form.cleaned_data.get('last_name')) > 0 and len(form.cleaned_data.get('gender')) > 0 and len(form.cleaned_data.get('phone')) == 11 or len(form.cleaned_data.get('phone')) == 0 or form.cleaned_data.get('phone').isdigit() and query_phone == 0:

            if len(form.cleaned_data.get('phone')) == 0:
                phonex = publisher.publisherprofile.phone
            else:
                phonex = phone_value

            user = form.save()
            user.refresh_from_db()
            user.publisherprofile.first_name = form.cleaned_data.get(
                'first_name')
            user.publisherprofile.middle_name = form.cleaned_data.get(
                'middle_name')
            user.publisherprofile.last_name = form.cleaned_data.get(
                'last_name')
            user.publisherprofile.username = form.cleaned_data.get(
                'first_name') + form.cleaned_data.get('username')
            user.publisherprofile.gender = form.cleaned_data.get('gender')
            user.publisherprofile.phone = phonex
            user.publisherprofile.pioneer = 0
            user.publisherprofile.appointment = 0
            user.publisherprofile.privilege = 0
            user.publisherprofile.parent_id = publisher.publisherprofile.username
            user.publisherprofile.status = 'Pending'
            user.publisherprofile.created_at = datetime.datetime.now()
            user.save()
            User.objects.filter(username=form.cleaned_data.get(
                'username')).update(phone="", password="")
            # username = form.cleaned_data.get('first_name') + form.cleaned_data.get('username')

            # username = form.cleaned_data.get('first_name') + form.cleaned_data.get('username') + get_random_string(length=6)
            # password = form.cleaned_data.get('password1')

            # user = authenticate(username=username, password=password)
            # login(response, user)
            # return HttpResponseRedirect("/sign-up-suc-redirect")
            form = AddFollowerForm()
            return redirect("/sign-up-suc")
        else:
            return redirect("/sign-up-suc")

    else:
        form = AddFollowerForm()

    followers = PublisherProfile.objects.filter(parent_id=publisher.publisherprofile.username).exclude(
        user_id=publisher.publisherprofile.user_id)
    # followers = PublisherProfile.objects.all()

    # update_follower_form = UpdateFollowerForm(response.POST)
    # if update_follower_form.is_valid():
    #     t = PublisherProfile.objects.get(id=1)

    # username2 = User.objects.all()
    return render(response, "main/sign-up-suc.html", {"publisher": publisher, "followers": followers, 'form': form})


def update_acct_pending_details(request):
    publisher = request.user
    if publisher.is_authenticated == False:
        return HttpResponseRedirect("/logout")

    if publisher.phone == "":
        return HttpResponseRedirect("/logout")

    if publisher.publisherprofile.status == "Active":
        return HttpResponseRedirect("/logout")

    if request.method == 'POST':

        userID = request.POST.get('userID')
        publisherModel = PublisherProfile.objects.get(user_id=userID)
        userModel = User.objects.get(id=userID)

        if request.POST.get("update_pen_details_submitButton"):
            phone_value = request.POST.get("phone")
            if len(phone_value) == 0:
                phonex = publisher.publisherprofile.phone
            else:
                phonex = phone_value

            query_phone = PublisherProfile.objects.filter(phone=phone_value).exclude(
                parent_id=publisher.publisherprofile.username).count()
            if len(request.POST.get("fname")) > 0 and len(request.POST.get("surname")) > 0 and len(request.POST.get("gender")) > 0 and len(request.POST.get("phone")) == 11 and request.POST.get("phone").isdigit() and query_phone == 0 and request.POST.get("phone") == request.POST.get("phone2"):
                publisherModel.first_name = request.POST.get("fname")
                publisherModel.middle_name = request.POST.get("mname")
                publisherModel.last_name = request.POST.get("surname")
                publisherModel.gender = request.POST.get("gender")
                publisherModel.email = request.POST.get("email")
                publisherModel.phone = phonex
                publisherModel.updated_at = datetime.datetime.now()
                # publisherModel.status = "Active"
                # publisherModel.groupName = 1
                # publisherModel.pioneer = 1
                # publisherModel.appointment = 1
                # publisherModel.privilege = 1
                publisherModel.save()

                userModel.first_name = request.POST.get("fname")
                userModel.last_name = request.POST.get("surname")
                userModel.email = request.POST.get("email")
                userModel.phone = phonex
                userModel.save()

            return redirect("/sign-up-suc/acct-pen-details/")

        elif request.POST.get("edit-follower-button"):
            phone_value = request.POST.get("secondary_phone")
            if len(phone_value) == 0:
                phonex = publisher.publisherprofile.phone
            else:
                phonex = phone_value

            query_phone = PublisherProfile.objects.filter(phone=phone_value).exclude(
                parent_id=publisher.publisherprofile.username).count()
            if len(request.POST.get("fname")) > 0 and len(request.POST.get("surname")) > 0 and len(request.POST.get("gender")) > 0 and len(request.POST.get("secondary_phone")) == 11 or len(request.POST.get("secondary_phone")) == 0 or request.POST.get("secondary_phone").isdigit() and query_phone == 0:
                publisherModel.first_name = request.POST.get("fname")
                publisherModel.middle_name = request.POST.get("mname")
                publisherModel.last_name = request.POST.get("surname")
                publisherModel.gender = request.POST.get("gender")
                publisherModel.phone = phonex
                publisherModel.updated_at = datetime.datetime.now()
                publisherModel.save()

                userModel.first_name = request.POST.get("fname")
                userModel.last_name = request.POST.get("surname")
                userModel.phone = ""
                userModel.save()

            return redirect("/sign-up-suc")

        elif request.POST.get("acct_pen_change_password_button"):

            if len(request.POST.get('password1')) > 0 and request.POST.get('password1') == request.POST.get('password2'):

                userModel.password = make_password(
                    request.POST.get("password1"))
                userModel.save()
                update_session_auth_hash(request, request.user)
                # publisher.is_authenticated
                # userModel = authenticate(username=userModel.phone, password=userModel.password)
                # login(request, userModel)

                publisherModel.updated_at = datetime.datetime.now()
                publisherModel.save()

            return redirect("/sign-up-suc/acct-pen-details/")
    else:
        return redirect("/sign-up-suc")


def check_follower_phone_exist(request):
    publisher = request.user

    if request.method == 'POST':
        phone = request.POST.get('phone')
        if request.POST.get('phone_check_sign_up'):
            if len(phone) > 0:
                query_phone = PublisherProfile.objects.filter(phone=phone)
                if(phone.isdigit()):
                    if query_phone.exists():
                        return HttpResponse("taken")
                    else:
                        return HttpResponse("not_taken")
                else:
                    return HttpResponse("not_number")
            else:
                return HttpResponse(0)
        else:
            if len(phone) > 0:
                query_phone = PublisherProfile.objects.filter(phone=phone).exclude(
                    parent_id=publisher.publisherprofile.username)
                if(phone.isdigit()):
                    if query_phone.exists():
                        return HttpResponse("taken")
                    else:
                        return HttpResponse("not_taken")
                else:
                    return HttpResponse("not_number")

            else:
                return HttpResponse(0)
    else:
        return redirect("/sign-up-suc")


def delete_follower(request):
    publisher = request.user

    if publisher.is_authenticated == False:
        return HttpResponseRedirect("/logout")

    if publisher.phone == "":
        return HttpResponseRedirect("/logout")

    if publisher.publisherprofile.status == "Active":
        return HttpResponseRedirect("/logout")

    if request.method == 'POST':
        id = request.POST.get('id')

        PublisherProfile.objects.get(user_id=id).delete()
        User.objects.get(id=id).delete()

        if len(id) == 0:
            return HttpResponse(0)
        else:
            return HttpResponse(1)
    else:
        return redirect("/sign-up-suc")


def acct_pen_details(response):
    publisher = response.user

    if publisher.is_authenticated == False:
        return HttpResponseRedirect("/logout")

    if publisher.phone == "":
        return HttpResponseRedirect("/logout")

    if publisher.publisherprofile.status == "Active":
        return HttpResponseRedirect("/logout")

    return render(response, "main/acct-pen-details.html", {"publisher": publisher})


def sign_up_suc_redirect(response):

    return redirect("/sign-up-suc")


def dashboard(response):
    publisher = response.user
    if publisher.is_authenticated == False:
        return HttpResponseRedirect("/logout")

    if publisher.publisherprofile.status == "Pending":
        return HttpResponseRedirect("/sign-up-suc")

    nav = "Dashboard"

    return render(response, "main/dashboard/dashboard.html", {"publisher": publisher, "nav": nav})


def month_year(response):
    publisher = response.user
    if publisher.is_authenticated == False:
        return HttpResponseRedirect("/logout")

    if publisher.publisherprofile.status == "Pending":
        return HttpResponseRedirect("/sign-up-suc")

    if not publisher.publisherprofile.privilege == "1":
        return HttpResponseRedirect("/logout")

    add_form = AddMonthYearForm(response.POST)
    if response.POST.get("add-month-year"):
        if add_form.is_valid():
            month = add_form.cleaned_data["month"]
            year = add_form.cleaned_data["year"]
            check_exist = MonthYear.objects.filter(
                month=month, year=year).count()
            if check_exist == 0 and year.isdigit() and len(month) > 0 and len(year) > 0:
                if add_form.cleaned_data["current"]:
                    MonthYear.objects.filter(
                        current=True).update(current=False)
                    l_current = True
                else:
                    l_current = False

                if add_form.cleaned_data["open"]:
                    open = True
                else:
                    open = False

                t = MonthYear(month=month, year=year,
                              current=l_current, open=open)
                t.save()
                return HttpResponseRedirect("/dashboard/month-year")
            else:
                return HttpResponseRedirect("/dashboard/month-year")
        else:
            AddMonthYearForm()

    if response.POST.get("edit-month-year"):
        id = response.POST.get('edit-id')
        Month_Year_model = MonthYear.objects.get(id=id)
        month = response.POST.get("month")
        year = response.POST.get("year")

        check_exist = MonthYear.objects.filter(
            month=month, year=year).exclude(id=id).count()
        if check_exist == 0 and year.isdigit() and len(month) > 0 and len(year) > 0:
            if response.POST.get("current"):
                MonthYear.objects.filter(current=True).update(current=False)
                l_current = True
            else:
                l_current = False

            if response.POST.get("open"):
                open = True
            else:
                open = False

            Month_Year_model.month = month
            Month_Year_model.year = year
            Month_Year_model.current = l_current
            Month_Year_model.open = open
            Month_Year_model.save()
            return HttpResponseRedirect("/dashboard/month-year")
        else:
            return HttpResponseRedirect("/dashboard/month-year")

    if response.POST.get('tag') == "delete":
        id = response.POST.get('id')

        MonthYear.objects.get(id=id).delete()

        if len(id) == 0:
            return HttpResponse(0)
        else:
            return HttpResponse(1)

    month_Year = MonthYear.objects.all().order_by('-id')

    return render(response, "main/dashboard/month-year.html", {"publisher": publisher, "add_form": add_form, "month_Year": month_Year})


def check_month_year(response):
    if response.POST.get('month_year_check'):
        month = response.POST.get('month')
        year = response.POST.get('year')
        if len(year) > 0:
            check_exist = MonthYear.objects.filter(
                month=month, year=year).count()
            if(year.isdigit()):
                if check_exist > 0:
                    return HttpResponse("taken")
                else:
                    return HttpResponse("not_taken")
            else:
                return HttpResponse("not_number")
        else:
            return HttpResponse(9)

    elif response.POST.get('month_year_check2'):
        month = response.POST.get('month')
        year = response.POST.get('year')
        id = response.POST.get('id')
        if len(year) > 0:
            check_exist = MonthYear.objects.filter(
                month=month, year=year).exclude(id=id).count()
            if(year.isdigit()):
                if check_exist > 0:
                    return HttpResponse("taken")
                else:
                    return HttpResponse("not_taken")
            else:
                return HttpResponse("not_number")
        else:
            return HttpResponse(9)


def waiting_room(response):

    publisher = response.user
    if publisher.is_authenticated == False:
        return HttpResponseRedirect("/logout")

    if publisher.publisherprofile.status == "Pending":
        return HttpResponseRedirect("/sign-up-suc")

    if not publisher.publisherprofile.privilege == "1":
        return HttpResponseRedirect("/logout")

    if response.POST.get('tag') == "delete":
        id = response.POST.get('id')

        PublisherProfile.objects.get(user_id=id).delete()

        if len(id) == 0:
            return HttpResponse(0)
        else:
            return HttpResponse(1)

    pending_publishers = PublisherProfile.objects.filter(status="Pending")

    for pending_publisher in pending_publishers:
        # p = PublisherProfile.objects.filter(parent_id=pending_publisherw.username).exclude(username="JudeAFA5F6513")
        p = pending_publisher

    parent_names = PublisherProfile.objects.all()
    count_followers = PublisherProfile.objects.filter(status="Pending")

    return render(response, "main/dashboard/waiting-room.html", {"publisher": publisher, "pending_publishers": pending_publishers, "parent_names": parent_names, "count_followers": count_followers})


def waiting_room_approve(response):
    publisher = response.user
    if publisher.is_authenticated == False:
        return HttpResponseRedirect("/logout")

    if publisher.publisherprofile.status == "Pending":
        return HttpResponseRedirect("/sign-up-suc")

    if response.method == 'POST':
        id = response.POST.get('unique_id')
        approve_model = PublisherProfile.objects.get(user_id=id)
        group = response.POST.get("group")

        if group == "Non-Publisher":
            pioneer = 0
            appointment = 0
        else:
            if response.POST.get("appoint"):
                appointment = response.POST.get("appoint")
            else:
                appointment = 0

            if response.POST.get("pioneer"):
                pioneer = 1
            else:
                pioneer = 0

        if len(group) > 0:
            approve_model.groupName = group
            approve_model.pioneer = pioneer
            approve_model.appointment = appointment
            approve_model.status = "Active"
            approve_model.updated_at = datetime.datetime.now()
            approve_model.save()
            return HttpResponseRedirect("/dashboard/waiting-room")
        else:
            return HttpResponseRedirect("/dashboard/waiting-room")


def submit_report_select_month_year(response):
    publisher = response.user
    if publisher.is_authenticated == False:
        return HttpResponseRedirect("/logout")

    if publisher.publisherprofile.status == "Pending":
        return HttpResponseRedirect("/sign-up-suc")

    if response.session.get('id', None):
        del response.session['id']

    if response.method == 'POST':
        month = response.POST.get('Month')
        year = response.POST.get('Year')
        check = MonthYear.objects.filter(month=month, year=year).first()
        if check:
            response.session['id'] = check.id
            return HttpResponseRedirect("/submit-report")
        else:
            return HttpResponseRedirect("/submit-report/select-month-year")

    current = MonthYear.objects.filter(current=True)
    open = MonthYear.objects.filter(open=True)

    return render(response, "main/report/select-report-month-year.html", {"publisher": publisher, "current": current, "open": open})


def submit_report(response):
    publisher = response.user
    if publisher.is_authenticated == False:
        return HttpResponseRedirect("/logout")

    if publisher.publisherprofile.status == "Pending":
        return HttpResponseRedirect("/sign-up-suc")

    if not response.session.get('id', None):
        return HttpResponseRedirect("/submit-report/select-month-year")

    month_year = MonthYear.objects.get(id=response.session['id'])

    my_publishers = PublisherProfile.objects.filter(
        parent_id=publisher.publisherprofile.username)
    my_publishers_add_form = PublisherProfile.objects.filter(
        parent_id=publisher.publisherprofile.username, status="Active").exclude(groupName="Non-Publisher")

    # reports = PublisherReport.objects.filter(month="January")

    return render(response, "main/report/submit-report.html", {"publisher": publisher, "month_year": month_year, "my_publishers": my_publishers, "my_publishers_add_form": my_publishers_add_form})


def submit_report_post(response):

    publisher = response.user
    if publisher.is_authenticated == False:
        return HttpResponseRedirect("/logout")

    if publisher.publisherprofile.status == "Pending":
        return HttpResponseRedirect("/sign-up-suc")

    if not response.session.get('id', None):
        return HttpResponseRedirect("/submit-report/select-month-year")

    month_year = MonthYear.objects.get(id=response.session['id'])
    if response.POST.get('add_report'):
        publisher_username = response.POST.get('select_publisher')

        if response.POST.get('return_visit').isdigit():
            return_visit = int(response.POST.get('return_visit'))
        else:
            return_visit = 0

        if response.POST.get('bible_study').isdigit():
            bible_study = int(response.POST.get('bible_study'))
        else:
            bible_study = 0

        if response.POST.get('placements').isdigit():
            placements = int(response.POST.get('placements'))
        else:
            placements = 0

        if response.POST.get('video').isdigit():
            video = int(response.POST.get('video'))
        else:
            video = 0

        hours = int(response.POST.get('hours'))
        comments = response.POST.get('comments')

        publisher_info = PublisherProfile.objects.filter(
            username=publisher_username).first()
        check_duplicate_report = PublisherReport.objects.filter(
            publisher_username=publisher_username, month_year_id=month_year.id).count()
        if check_duplicate_report == 0:
            g = PublisherReport(publisher_username=publisher_info.username,
                                group=publisher_info.groupName,
                                pioneer=publisher_info.pioneer,
                                appointment=publisher_info.appointment,
                                month=month_year.month,
                                year=month_year.year,
                                month_year_id=month_year.id,
                                placements=placements,
                                video_showings=video,
                                hours=hours,
                                return_visits=return_visit,
                                bible_studies=bible_study,
                                comments=comments,
                                created_at=datetime.datetime.now())

            if int(bible_study) <= int(return_visit):
                g.save()

        # PublisherReport.objects.all().delete()
        return HttpResponseRedirect("/submit-report")

    if response.POST.get('tag') == "delete":
        id = response.POST.get('id')

        PublisherReport.objects.get(
            publisher_username=id, month_year_id=month_year.id).delete()

        if len(id) == 0:
            return HttpResponse(0)
        else:
            return HttpResponse(1)


def submit_report_edit(response):

    publisher = response.user
    if publisher.is_authenticated == False:
        return HttpResponseRedirect("/logout")

    if publisher.publisherprofile.status == "Pending":
        return HttpResponseRedirect("/sign-up-suc")

    if not response.session.get('id', None):
        return HttpResponseRedirect("/submit-report/select-month-year")

    if response.POST.get('post_edit') == "1":

        id = response.POST.get('report_id')
        edit_report = PublisherReport.objects.get(id=id)

        if response.POST.get('return_visit').isdigit():
            return_visit = int(response.POST.get('return_visit'))
        else:
            return_visit = 0

        if response.POST.get('bible_study').isdigit():
            bible_study = int(response.POST.get('bible_study'))
        else:
            bible_study = 0

        if response.POST.get('placements').isdigit():
            placements = int(response.POST.get('placements'))
        else:
            placements = 0

        if response.POST.get('video').isdigit():
            video = int(response.POST.get('video'))
        else:
            video = 0

        hours = int(response.POST.get('hours'))
        comments = response.POST.get('comments')

        if int(bible_study) <= int(return_visit):
            edit_report.placements = placements
            edit_report.video_showings = video
            edit_report.hours = hours
            edit_report.return_visits = return_visit
            edit_report.bible_studies = bible_study
            edit_report.comments = comments
            edit_report.updated_at = datetime.datetime.now()
            edit_report.save()
        return HttpResponseRedirect("/submit-report")
    else:
        return HttpResponseRedirect("/submit-report")


def all_publishers(response):
    publisher = response.user
    if publisher.is_authenticated == False:
        return HttpResponseRedirect("/logout")

    if publisher.publisherprofile.status == "Pending":
        return HttpResponseRedirect("/sign-up-suc")

    all_publishers = PublisherProfile.objects.filter(
        status="Active").exclude(groupName="Non-Publisher")
    count_active_publishers = PublisherProfile.objects.filter(
        status="Active").exclude(groupName="Non-Publisher").count()

    nav = "Publishers"
    return render(response, "main/all-publishers/all-publishers.html", {"publisher": publisher, "all_publishers": all_publishers, "count_active_publishers": count_active_publishers, "nav": nav})


def non_publishers(response):
    publisher = response.user
    if publisher.is_authenticated == False:
        return HttpResponseRedirect("/logout")

    if publisher.publisherprofile.status == "Pending":
        return HttpResponseRedirect("/sign-up-suc")

    if not publisher.publisherprofile.privilege == "1":
        return HttpResponseRedirect("/logout")

    non_publishers = PublisherProfile.objects.filter(groupName="Non-Publisher")
    count_non_publishers = PublisherProfile.objects.filter(
        groupName="Non-Publisher").count()

    nav = "Publishers"
    return render(response, "main/all-publishers/non-publishers.html", {"publisher": publisher, "non_publishers": non_publishers, "count_non_publishers": count_non_publishers, "nav": nav})


def deactivated_publishers(response):
    publisher = response.user
    if publisher.is_authenticated == False:
        return HttpResponseRedirect("/logout")

    if publisher.publisherprofile.status == "Pending":
        return HttpResponseRedirect("/sign-up-suc")

    if not publisher.publisherprofile.privilege == "1":
        return HttpResponseRedirect("/logout")

    deactivated_publishers = PublisherProfile.objects.filter(
        status="deactivated")
    count_deactivated_publishers = PublisherProfile.objects.filter(
        status="deactivated").count()

    nav = "Publishers"
    return render(response, "main/all-publishers/deactivated-publishers.html", {"publisher": publisher, "deactivated_publishers": deactivated_publishers, "count_deactivated_publishers": count_deactivated_publishers, "nav": nav})


def publisher_profile(response, username):
    publisher = response.user
    if publisher.is_authenticated == False:
        return HttpResponseRedirect("/logout")

    if publisher.publisherprofile.status == "Pending":
        return HttpResponseRedirect("/sign-up-suc")
    
    if len(username) == 0:
        return HttpResponseRedirect("/publishers/all-active-publishers")
    try:
        profile = PublisherProfile.objects.get(username=username)
    except (PublisherProfile.DoesNotExist):
        return HttpResponseRedirect("/publishers/all-active-publishers")
    
    

    nav = publisher.first_name
    return render(response, "main/publisher-profile/publisher-profile.html", {"publisher": publisher, "profile": profile, "nav": nav})


def choose_category(response):
    publisher = response.user
    if publisher.is_authenticated == False:
        return HttpResponseRedirect("/logout")

    if publisher.publisherprofile.status == "Pending":
        return HttpResponseRedirect("/sign-up-suc")

    if not publisher.publisherprofile.privilege == "1":
        return HttpResponseRedirect("/logout")

    nav = "Publishers"
    err = ""

    if response.session.get('publishers_report_month_year_id', None):
        del response.session['publishers_report_month_year_id']
        del response.session['publishers_report_category']

    current = MonthYear.objects.get(current=True)
    all_month = MonthYear.objects.raw(
        'SELECT * FROM main_monthyear GROUP BY month ORDER BY id DESC')
    all_year = MonthYear.objects.raw(
        'SELECT * FROM main_monthyear GROUP BY year ORDER BY id DESC')

    if response.method == 'POST':
        month = response.POST.get('Month')
        year = response.POST.get('Year')
        Category = response.POST.get('Category')
        check = MonthYear.objects.filter(month=month, year=year).first()
        if Category == "one":
            report = PublisherReport.objects.filter(month_year_id=check.id)
        elif Category == "two":
            report = PublisherReport.objects.filter(
                month_year_id=check.id, pioneer="1")
        elif Category == "three":
            report = PublisherReport.objects.filter(
                month_year_id=check.id, pioneer="2")
        elif Category == "four":
            report = PublisherReport.objects.filter(
                month_year_id=check.id, appointment="1")
        elif Category == "five":
            report = PublisherReport.objects.filter(
                month_year_id=check.id, appointment="2")

        if report.count() > 0:
            response.session['publishers_report_month_year_id'] = check.id
            response.session['publishers_report_category'] = Category
            return HttpResponseRedirect("/publishers/monthly-reports")
        else:
            # check.id = ""
            err = "No Record Found"

    return render(response, "main/all-publishers/choose-category.html", {"publisher": publisher, "nav": nav, "current": current, "all_month": all_month, "all_year": all_year, "err": err})


def monthly_report(response):
    publisher = response.user
    if publisher.is_authenticated == False:
        return HttpResponseRedirect("/logout")

    if publisher.publisherprofile.status == "Pending":
        return HttpResponseRedirect("/sign-up-suc")

    if not publisher.publisherprofile.privilege == "1":
        return HttpResponseRedirect("/logout")

    if not response.session.get('publishers_report_month_year_id', None):
        return HttpResponseRedirect("/publishers/choose-category")

    month_year = MonthYear.objects.get(
        id=response.session['publishers_report_month_year_id'])
    Category = response.session['publishers_report_category']

    if Category == "one":
        category = "All Publishers"
        reports = PublisherReport.objects.filter(month_year_id=month_year.id)
        count_publishers = PublisherProfile.objects.filter(
            status="Active").exclude(groupName="Non-Publisher").count()
    elif Category == "two":
        category = "Regular Pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, pioneer="1")
        count_publishers = PublisherProfile.objects.filter(
            status="Active", pioneer="1").exclude(groupName="Non-Publisher").count()
    elif Category == "three":
        category = "Auxiliary Pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, pioneer="2")
        count_publishers = PublisherProfile.objects.filter(
            status="Active", pioneer="2").exclude(groupName="Non-Publisher").count()
    elif Category == "four":
        category = "Elders"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, appointment="1")
        count_publishers = PublisherProfile.objects.filter(
            status="Active", appointment="1").exclude(groupName="Non-Publisher").count()
    elif Category == "five":
        category = "Ministerial Servants"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, appointment="2")
        count_publishers = PublisherProfile.objects.filter(
            status="Active", appointment="2").exclude(groupName="Non-Publisher").count()

    count_report = reports.count()

    if Category == "one":
        tag = "Current Number of Active Publishers"
    else:
        tag = "Current Number of "+category

    if response.POST.get('tag') == "delete":
        idr = response.POST.get('id')

        PublisherReport.objects.get(id=idr).delete()

        if len(idr) == 0:
            return HttpResponse(0)
        else:
            return HttpResponse(1)

    nav = "Publishers"
    return render(response, "main/all-publishers/all-publishers-report.html", {"publisher": publisher, "month_year": month_year, "category": category, "reports": reports, "count_publishers": count_publishers, "count_report": count_report, "tag": tag, "nav": nav})


# def monthly_report_delete(response):

#     if response.POST.get('tag') == "delete":
#         id = response.POST.get('id')

#         PublisherReport.objects.filter(id=id).delete()

#         if len(id) == 0:
#             return HttpResponse(0)
#         else:
#             return HttpResponse(1)

def monthly_report_stats(response):
    publisher = response.user
    if publisher.is_authenticated == False:
        return HttpResponseRedirect("/logout")

    if publisher.publisherprofile.status == "Pending":
        return HttpResponseRedirect("/sign-up-suc")

    if not publisher.publisherprofile.privilege == "1":
        return HttpResponseRedirect("/logout")

    if not response.session.get('publishers_report_month_year_id', None):
        return HttpResponseRedirect("/publishers/choose-category")

    month_year = MonthYear.objects.get(
        id=response.session['publishers_report_month_year_id'])
    Category = response.session['publishers_report_category']

    if Category == "one":
        category = "All Publishers"
        reports = PublisherReport.objects.filter(month_year_id=month_year.id)

    elif Category == "two":
        category = "Regular Pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, pioneer="1")

    elif Category == "three":
        category = "Auxiliary Pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, pioneer="2")

    elif Category == "four":
        category = "Elders"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, appointment="1")

    elif Category == "five":
        category = "Ministerial Servants"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, appointment="2")

    count_report = reports.count()
    total_hours = reports.aggregate(
        total=Sum('hours'))['total']
    total_placements = reports.aggregate(
        total=Sum('placements'))['total']
    total_return_visits = reports.aggregate(
        total=Sum('return_visits'))['total']
    total_bible_studies = reports.aggregate(
        total=Sum('bible_studies'))['total']
    total_video_showings = reports.aggregate(
        total=Sum('video_showings'))['total']

    avg_hours = reports.aggregate(
        average=Avg('hours'))['average']
    avg_placements = reports.aggregate(
        average=Avg('placements'))['average']
    avg_return_visits = reports.aggregate(
        average=Avg('return_visits'))['average']
    avg_bible_studies = reports.aggregate(
        average=Avg('bible_studies'))['average']
    avg_video_showings = reports.aggregate(
        average=Avg('video_showings'))['average']

    # avg_hours = total_hours/count_report

    nav = "Publishers"
    return render(response, "main/all-publishers/all-publishers-report-stats.html", {"publisher": publisher, "month_year": month_year, "category": category, "reports": reports, "count_report": count_report,
                                                                                     "total_hours": total_hours, "total_placements": total_placements, "total_return_visits": total_return_visits,
                                                                                     "total_bible_studies": total_bible_studies, "total_video_showings": total_video_showings,
                                                                                     "avg_hours": avg_hours, "avg_placements": avg_placements, "avg_return_visits": avg_return_visits,
                                                                                     "avg_bible_studies": avg_bible_studies, "avg_video_showings": avg_video_showings,
                                                                                     "nav": nav})


def monthly_pending_report(response):
    publisher = response.user
    if publisher.is_authenticated == False:
        return HttpResponseRedirect("/logout")

    if publisher.publisherprofile.status == "Pending":
        return HttpResponseRedirect("/sign-up-suc")

    if not publisher.publisherprofile.privilege == "1":
        return HttpResponseRedirect("/logout")

    if not response.session.get('publishers_report_month_year_id', None):
        return HttpResponseRedirect("/publishers/choose-category")

    month_year = MonthYear.objects.get(
        id=response.session['publishers_report_month_year_id'])
    Category = response.session['publishers_report_category']

    if Category == "one":
        category = "All Publishers"
        reports = PublisherReport.objects.filter(month_year_id=month_year.id)
        publishers = PublisherProfile.objects.filter(
            status="Active").exclude(groupName="Non-Publisher")
    elif Category == "two":
        category = "Regular Pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, pioneer="1")
        publishers = PublisherProfile.objects.filter(
            status="Active", pioneer="1").exclude(groupName="Non-Publisher")
    elif Category == "three":
        category = "Auxiliary Pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, pioneer="2")
        publishers = PublisherProfile.objects.filter(
            status="Active", pioneer="2").exclude(groupName="Non-Publisher")
    elif Category == "four":
        category = "Elders"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, appointment="1")
        publishers = PublisherProfile.objects.filter(
            status="Active", appointment="1").exclude(groupName="Non-Publisher")
    elif Category == "five":
        category = "Ministerial Servants"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, appointment="2")
        publishers = PublisherProfile.objects.filter(
            status="Active", appointment="2").exclude(groupName="Non-Publisher")

    count_publishers = publishers.count()
    count_report = reports.count()

    count_pending_reports = count_publishers - count_report

    if count_pending_reports > 1:
        pending_reports_tag = 'Pending reports'
    elif count_pending_reports == 1:
        pending_reports_tag = 'Pending report'
    else:
        pending_reports_tag = 'No Pending report'
        count_pending_reports = ''

    nav = "Publishers"
    return render(response, "main/all-publishers/pending-reports.html", {"publisher": publisher, "month_year": month_year, "category": category, "reports": reports, "publishers": publishers, "count_pending_reports": count_pending_reports, "pending_reports_tag": pending_reports_tag, "nav": nav})


def export_import_monthly_report(response):
    publisher = response.user
    if publisher.is_authenticated == False:
        return HttpResponseRedirect("/logout")

    if publisher.publisherprofile.status == "Pending":
        return HttpResponseRedirect("/sign-up-suc")

    if not publisher.publisherprofile.privilege == "1":
        return HttpResponseRedirect("/logout")

    if not response.session.get('publishers_report_month_year_id', None):
        return HttpResponseRedirect("/publishers/choose-category")

    month_year = MonthYear.objects.get(
        id=response.session['publishers_report_month_year_id'])
    Category = response.session['publishers_report_category']

    if Category == "one":
        category = "All Publishers"
        reports = PublisherReport.objects.filter(month_year_id=month_year.id)
        publishers = PublisherProfile.objects.filter(
            status="Active").exclude(groupName="Non-Publisher")
    elif Category == "two":
        category = "Regular Pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, pioneer="1")
        publishers = PublisherProfile.objects.filter(
            status="Active", pioneer="1").exclude(groupName="Non-Publisher")
    elif Category == "three":
        category = "Auxiliary Pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, pioneer="2")
        publishers = PublisherProfile.objects.filter(
            status="Active", pioneer="2").exclude(groupName="Non-Publisher")
    elif Category == "four":
        category = "Elders"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, appointment="1")
        publishers = PublisherProfile.objects.filter(
            status="Active", appointment="1").exclude(groupName="Non-Publisher")
    elif Category == "five":
        category = "Ministerial Servants"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, appointment="2")
        publishers = PublisherProfile.objects.filter(
            status="Active", appointment="2").exclude(groupName="Non-Publisher")

    count_publishers = publishers.count()
    count_report = reports.count()

    count_pending_reports = count_publishers - count_report

    if count_pending_reports > 1:
        pending_reports_tag = 'Pending reports'
    elif count_pending_reports == 1:
        pending_reports_tag = 'Pending report'
    else:
        pending_reports_tag = 'No Pending report'
        count_pending_reports = ''

    tests = PublisherReport.objects.filter(
        month_year_id="2")

    nav = "Publishers"
    return render(response, "main/all-publishers/export-import-reports.html", {"tests": tests, "publisher": publisher, "month_year": month_year, "category": category, "reports": reports, "publishers": publishers, "count_pending_reports": count_pending_reports, "pending_reports_tag": pending_reports_tag, "nav": nav})


def export_monthly_report_csv(request):
    if not request.session.get('publishers_report_month_year_id', None):
        return HttpResponseRedirect("/publishers/choose-category")

    month_year = MonthYear.objects.get(
        id=request.session['publishers_report_month_year_id'])
    Category = request.session['publishers_report_category']

    if Category == "one":
        category_tag = "all_publishers"
        reports = PublisherReport.objects.filter(month_year_id=month_year.id).values_list(
            'publisher_username', 'month', 'year', 'pioneer', 'appointment', 'placements', 'video_showings', 'hours', 'return_visits', 'bible_studies', 'comments')
    elif Category == "two":
        category_tag = "regular_pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, pioneer="1").values_list('publisher_username', 'month', 'year', 'pioneer', 'appointment', 'placements', 'video_showings', 'hours', 'return_visits', 'bible_studies', 'comments')
    elif Category == "three":
        category_tag = "auxiliary_pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, pioneer="2").values_list('publisher_username', 'month', 'year', 'pioneer', 'appointment', 'placements', 'video_showings', 'hours', 'return_visits', 'bible_studies', 'comments')
    elif Category == "four":
        category_tag = "elders"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, appointment="1").values_list('publisher_username', 'month', 'year', 'pioneer', 'appointment', 'placements', 'video_showings', 'hours', 'return_visits', 'bible_studies', 'comments')
    elif Category == "five":
        category_tag = "ministerial_servants"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, appointment="2").values_list('publisher_username', 'month', 'year', 'pioneer', 'appointment', 'placements', 'video_showings', 'hours', 'return_visits', 'bible_studies', 'comments')

    filename = category_tag+'_report_for_'+month_year.month+'_'+month_year.year

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="'+filename+'.csv"'

    writer = csv.writer(response)
    writer.writerow(['First Name', 'Middle Name', 'Last Name', 'Username', 'Month', 'Year', 'Pioneer', 'Appointment', 'Placements',
                    'Video Showings', 'Hours', 'Return Visits', 'Bible Studies', 'Comments'])

    for row in reports:
        publisher = PublisherProfile.objects.get(
                username=row[0])

        if row[3] == "1":
            pioneer = "R"
        else:
            pioneer = "A"

        if row[4] == "1":
            appointment = "E"
        elif row[4] == "2":
            appointment = "MS"
        else:
            appointment = ""

        writer.writerow([publisher.first_name, publisher.first_name, publisher.last_name, row[0],
                        row[1], row[2], pioneer, appointment, row[5], row[6], row[7], row[8], row[9], row[10]])

    return response


def export_monthly_report_xls(request):

    if not request.session.get('publishers_report_month_year_id', None):
        return HttpResponseRedirect("/publishers/choose-category")

    month_year = MonthYear.objects.get(
        id=request.session['publishers_report_month_year_id'])
    Category = request.session['publishers_report_category']

    if Category == "one":
        category_tag = "all_publishers"
        reports = PublisherReport.objects.filter(month_year_id=month_year.id).values_list(
            'publisher_username', 'month', 'year', 'pioneer', 'appointment', 'placements', 'video_showings', 'hours', 'return_visits', 'bible_studies', 'comments')
    elif Category == "two":
        category_tag = "regular_pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, pioneer="1").values_list('publisher_username', 'month', 'year', 'pioneer', 'appointment', 'placements', 'video_showings', 'hours', 'return_visits', 'bible_studies', 'comments')
    elif Category == "three":
        category_tag = "auxiliary_pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, pioneer="2").values_list('publisher_username', 'month', 'year', 'pioneer', 'appointment', 'placements', 'video_showings', 'hours', 'return_visits', 'bible_studies', 'comments')
    elif Category == "four":
        category_tag = "elders"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, appointment="1").values_list('publisher_username', 'month', 'year', 'pioneer', 'appointment', 'placements', 'video_showings', 'hours', 'return_visits', 'bible_studies', 'comments')
    elif Category == "five":
        category_tag = "ministerial_servants"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, appointment="2").values_list('publisher_username', 'month', 'year', 'pioneer', 'appointment', 'placements', 'video_showings', 'hours', 'return_visits', 'bible_studies', 'comments')

    filename = category_tag+'_report_for_'+month_year.month+'_'+month_year.year

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="'+filename+'.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Reports')

    # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['FirstName', 'Middle Name', 'Last Name', 'Username', 'Month', 'Year', 'Pioneer', 'Appointment', 'Placements',
               'Video Showings', 'Hours', 'Return Visits', 'Bible Studies', 'Comments']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    # rows = User.objects.all().values_list('username', 'first_name', 'last_name', 'email')
    for row in reports:
        # row_num = 1
        row_num += 1

        #  convert tuple to list to update column
        my_list = list(row)
        if my_list[3] == "1":
            my_list[3] = "R"
        else:
            my_list[3] = "A"

        if my_list[4] == "1":
            my_list[4] = "E"
        elif my_list[4] == "2":
            my_list[4] = "MS"
        else:
            my_list[4] = ""

        row = tuple(my_list)
        publisher = PublisherProfile.objects.get(
                username=row[0])

        ws.write(row_num, 0, publisher.first_name, font_style)
        ws.write(row_num, 1, publisher.middle_name, font_style)
        ws.write(row_num, 2, publisher.last_name, font_style)
        ws.write(row_num, 3, row[0], font_style)
        ws.write(row_num, 4, row[1], font_style)
        ws.write(row_num, 5, row[2], font_style)
        ws.write(row_num, 6, row[3], font_style)
        ws.write(row_num, 7, row[4], font_style)
        ws.write(row_num, 8, row[5], font_style)
        ws.write(row_num, 9, row[6], font_style)
        ws.write(row_num, 10, row[7], font_style)
        ws.write(row_num, 11, row[8], font_style)
        ws.write(row_num, 12, row[9], font_style)
        ws.write(row_num, 13, row[10], font_style)

    wb.save(response)
    return response


def import_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES["excel_file"]
        
        
        
        # Check if the file is an Excel file (XLSX)
        if excel_file.name.endswith('.xlsx'):
            try:
                wb = load_workbook(excel_file)
                sheet = wb.active

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    
                    month_year = MonthYear.objects.get(
                    month=row[4], year=row[5])
                    
                    report = PublisherReport.objects.filter(
                    month_year_id=month_year.id, publisher_username=row[3])
                    
                    publisher_exist = PublisherProfile.objects.filter(username=row[3]).exists()
                    
                    if publisher_exist:
                        if report.count() == 0:
                        
                            # Your data processing logic here
                            PublisherReport.objects.create(
                                publisher_username=row[3],
                                month=row[4],
                                year=row[5],
                                month_year_id=month_year.id,
                                pioneer=row[6],
                                appointment=row[7],
                                placements="2",
                                video_showings="2",
                                hours="2",
                                return_visits="2",
                                bible_studies="2",
                                comments=row[13],
                                created_at=datetime.datetime.now()
                            )

                messages.success(request, 'XLSX file uploaded successfully!')
            except Exception as e:
                messages.error(request, f'Error processing Excel file: {str(e)}')

         # Check if the file is an Excel file (XLS)
        elif excel_file.name.endswith('.xls'):
            try:
                # Read the XLS file using xlrd
                book = xlrd.open_workbook(file_contents=excel_file.read())
                sheet = book.sheet_by_index(0)

                for row_num in range(1, sheet.nrows):
                    row = sheet.row_values(row_num)
                    
                    month_year = MonthYear.objects.get(
                    month=row[4], year=row[5])
                    
                    report = PublisherReport.objects.filter(
                    month_year_id=month_year.id, publisher_username=row[3])
                    
                    publisher_exist = PublisherProfile.objects.filter(username=row[3]).exists()
                    
                    if publisher_exist:
                        if report.count() == 0:
                    
                            # Your data processing logic here
                            PublisherReport.objects.create(
                                publisher_username=row[3],
                                month=row[4],
                                year=row[5],
                                month_year_id=month_year.id,
                                pioneer=row[6],
                                appointment=row[7],
                                placements="2",
                                video_showings="2",
                                hours="2",
                                return_visits="2",
                                bible_studies="2",
                                comments=row[13],
                                created_at=datetime.datetime.now()
                            )

                messages.success(request, 'XLS file uploaded successfully!')
            except Exception as e:
                messages.error(request, f'Error processing XLS file: {str(e)}')

        # Check if the file is a CSV file
        elif excel_file.name.endswith('.csv'):
            try:
                # Read the CSV file
                decoded_file = TextIOWrapper(excel_file.file, encoding='utf-8')
                reader = csv.reader(decoded_file)
                # Skip the header row
                next(reader, None)

                for row in reader:
                    
                    month_year = MonthYear.objects.get(
                    month=row[4], year=row[5])
                    
                    report = PublisherReport.objects.filter(
                    month_year_id=month_year.id, publisher_username=row[3])
                    
                    publisher_exist = PublisherProfile.objects.filter(username=row[3]).exists()
                    
                    if publisher_exist:
                        if report.count() == 0:
                        
                            # Your data processing logic here
                            PublisherReport.objects.create(
                                publisher_username=row[3],
                                month=row[4],
                                year=row[5],
                                month_year_id=month_year.id,
                                pioneer=row[6],
                                appointment=row[7],
                                placements="2",
                                video_showings="2",
                                hours="2",
                                return_visits="2",
                                bible_studies="2",
                                comments=row[13],
                                created_at=datetime.datetime.now()
                            )

                messages.success(request, 'CSV file uploaded successfully!')
            except Exception as e:
                messages.error(request, f'Error processing CSV file: {str(e)}')

        else:
            messages.error(request, 'Invalid file format. Please upload an Excel (XLSX) or CSV file.')

    else:
        messages.error(request, 'Invalid request. Please upload a file.')

    return redirect('/publishers/export-import-monthly-reports')
       
def delete_reports(request):
    publisher=request.user

    if publisher.is_authenticated == False:
        return HttpResponseRedirect("/logout")

    if publisher.publisherprofile.status == "Pending":
        return HttpResponseRedirect("/sign-up-suc")

    if not publisher.publisherprofile.privilege == "1":
        return HttpResponseRedirect("/logout")

    if not request.session.get('publishers_report_month_year_id', None):
        return HttpResponseRedirect("/publishers/choose-category")


    if request.method == 'POST':
        # id = request.POST.get('id')

        PublisherReport.objects.filter(month_year_id = "2").delete()
        # User.objects.get(id=id).delete()

        # if len(id) == 0:
        #     return HttpResponse(0)
        # else:
        return HttpResponse(1)
    else:
        return redirect("/sign-up-suc")
