from django import template
import uuid
from calendar import month
from cmath import nan
from multiprocessing import cpu_count
from queue import Empty
from django.http import HttpResponse, HttpResponseRedirect, Http404
from django.contrib import messages
from django.contrib.auth import login, authenticate
from django.shortcuts import render, redirect
from ..models import PublisherProfile, User, MonthYear, PublisherReport
from django.utils.crypto import get_random_string
import datetime
from django.contrib.auth.hashers import make_password
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth import get_user_model
from django.db.models import Sum
from django.db.models import Avg, Count
import csv
from django.views import View
# pip install xlwt first
import xlwt

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font

import csv
from io import TextIOWrapper
import xlrd
from django.contrib import messages

from ..decorators import authentication_required, admin_privilege, check_month_year



@authentication_required
def all_publishers(response):
    publisher = response.user

    # if publisher.publisherprofile.status == "Pending":
    #     return HttpResponseRedirect("/sign-up-suc")

    all_publishers = PublisherProfile.objects.filter(
        status="Active").exclude(groupName="Non-Publisher")
    count_active_publishers = PublisherProfile.objects.filter(
        status="Active").exclude(groupName="Non-Publisher").count()

    nav = "Publishers"
    return render(response, "main/all-publishers/all-publishers.html", {"publisher": publisher, "all_publishers": all_publishers, "count_active_publishers": count_active_publishers, "nav": nav})

@authentication_required
@admin_privilege
def non_publishers(response):
    publisher = response.user
  
    # if publisher.publisherprofile.status == "Pending":
    #     return HttpResponseRedirect("/sign-up-suc")

    # if not publisher.publisherprofile.privilege == "1":
    #     return HttpResponseRedirect("/logout")

    non_publishers = PublisherProfile.objects.filter(groupName="Non-Publisher")
    count_non_publishers = PublisherProfile.objects.filter(
        groupName="Non-Publisher").count()

    nav = "Publishers"
    return render(response, "main/all-publishers/non-publishers.html", {"publisher": publisher, "non_publishers": non_publishers, "count_non_publishers": count_non_publishers, "nav": nav})

@authentication_required
@admin_privilege
def deactivated_publishers(response):
    publisher = response.user
   
    deactivated_publishers = PublisherProfile.objects.filter(
        status="deactivated")
    count_deactivated_publishers = PublisherProfile.objects.filter(
        status="deactivated").count()

    nav = "Publishers"
    return render(response, "main/all-publishers/deactivated-publishers.html", {"publisher": publisher, "deactivated_publishers": deactivated_publishers, "count_deactivated_publishers": count_deactivated_publishers, "nav": nav})


@authentication_required
@admin_privilege
@check_month_year
def choose_category(response):
    publisher = response.user
  
    nav = "Publishers"
    err = ""

    if response.session.get('publishers_report_month_year_id', None):
        del response.session['publishers_report_month_year_id']
        del response.session['publishers_report_category']

    current = MonthYear.objects.get(current=True)
    all_month = MonthYear.objects.raw(
        'SELECT * FROM main_monthyear GROUP BY month ORDER BY id DESC')
    all_year = MonthYear.objects.raw(
        'SELECT * FROM main_monthyear GROUP BY year ORDER BY id DESC')

    if response.method == 'POST':
        month = response.POST.get('Month')
        year = response.POST.get('Year')
        Category = response.POST.get('Category')
        check = MonthYear.objects.filter(month=month, year=year).first()
        if check is None:
            # raise Http404(f"{month} {year} not found")
            messages.error(response, f"ERROR: {month} {year} not registered")
            return HttpResponseRedirect("/publishers/choose-category")
            
        if Category == "one":
            report = PublisherReport.objects.filter(month_year_id=check.id)
        elif Category == "two":
            report = PublisherReport.objects.filter(
                month_year_id=check.id, pioneer="1")
        elif Category == "three":
            report = PublisherReport.objects.filter(
                month_year_id=check.id, pioneer="2")
        elif Category == "four":
            report = PublisherReport.objects.filter(
                month_year_id=check.id, appointment="1")
        elif Category == "five":
            report = PublisherReport.objects.filter(
                month_year_id=check.id, appointment="2")

        if report.count() > 0:
            response.session['publishers_report_month_year_id'] = check.id
            response.session['publishers_report_category'] = Category
            return HttpResponseRedirect("/publishers/monthly-reports")
        else:
            messages.error(response, f"ERROR: No record found under {month} {year}")

    return render(response, "main/all-publishers/choose-category.html", {"publisher": publisher, "nav": nav, "current": current, "all_month": all_month, "all_year": all_year, "err": err})

@authentication_required
@admin_privilege
@check_month_year
def monthly_report(response):
    publisher = response.user
   
    if not response.session.get('publishers_report_month_year_id', None):
        return HttpResponseRedirect("/publishers/choose-category")

    month_year = MonthYear.objects.get(
        id=response.session['publishers_report_month_year_id'])
    Category = response.session['publishers_report_category']

    if Category == "one":
        category = "All Publishers"
        reports = PublisherReport.objects.filter(month_year_id=month_year.id)
        count_publishers = PublisherProfile.objects.filter(
            status="Active").exclude(groupName="Non-Publisher").count()
    elif Category == "two":
        category = "Regular Pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, pioneer="1")
        count_publishers = PublisherProfile.objects.filter(
            status="Active", pioneer="1").exclude(groupName="Non-Publisher").count()
    elif Category == "three":
        category = "Auxiliary Pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, pioneer="2")
        count_publishers = PublisherProfile.objects.filter(
            status="Active", pioneer="2").exclude(groupName="Non-Publisher").count()
    elif Category == "four":
        category = "Elders"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, appointment="1")
        count_publishers = PublisherProfile.objects.filter(
            status="Active", appointment="1").exclude(groupName="Non-Publisher").count()
    elif Category == "five":
        category = "Ministerial Servants"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, appointment="2")
        count_publishers = PublisherProfile.objects.filter(
            status="Active", appointment="2").exclude(groupName="Non-Publisher").count()

    count_report = reports.count()
    
    if reports.count() == 0:
        return HttpResponseRedirect("/publishers/choose-category")

    if Category == "one":
        tag = "Current Number of Active Publishers"
    else:
        tag = "Current Number of "+category

    if response.POST.get('tag') == "delete":
        idr = response.POST.get('id')

        PublisherReport.objects.get(id=idr).delete()

        if len(idr) == 0:
            return HttpResponse(0)
        else:
            return HttpResponse(1)

    nav = "Publishers"
    return render(response, "main/all-publishers/all-publishers-report.html", {"publisher": publisher, "month_year": month_year, "category": category, "reports": reports, "count_publishers": count_publishers, "count_report": count_report, "tag": tag, "nav": nav})


# def monthly_report_delete(response):

#     if response.POST.get('tag') == "delete":
#         id = response.POST.get('id')

#         PublisherReport.objects.filter(id=id).delete()

#         if len(id) == 0:
#             return HttpResponse(0)
#         else:
#             return HttpResponse(1)

@authentication_required
@admin_privilege
@check_month_year
def monthly_report_stats(response):
    publisher = response.user

    if not response.session.get('publishers_report_month_year_id', None):
        return HttpResponseRedirect("/publishers/choose-category")

    month_year = MonthYear.objects.get(
        id=response.session['publishers_report_month_year_id'])
    Category = response.session['publishers_report_category']

    if Category == "one":
        category = "All Publishers"
        reports = PublisherReport.objects.filter(month_year_id=month_year.id)

    elif Category == "two":
        category = "Regular Pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, pioneer="1")

    elif Category == "three":
        category = "Auxiliary Pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, pioneer="2")

    elif Category == "four":
        category = "Elders"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, appointment="1")

    elif Category == "five":
        category = "Ministerial Servants"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, appointment="2")
        
        
    if reports.count() == 0:
        return HttpResponseRedirect("/publishers/choose-category")

    count_report = reports.count()
    total_hours = reports.aggregate(
        total=Sum('hours'))['total']
    total_placements = reports.aggregate(
        total=Sum('placements'))['total']
    total_return_visits = reports.aggregate(
        total=Sum('return_visits'))['total']
    total_bible_studies = reports.aggregate(
        total=Sum('bible_studies'))['total']
    total_video_showings = reports.aggregate(
        total=Sum('video_showings'))['total']

    avg_hours = reports.aggregate(
        average=Avg('hours'))['average']
    avg_placements = reports.aggregate(
        average=Avg('placements'))['average']
    avg_return_visits = reports.aggregate(
        average=Avg('return_visits'))['average']
    avg_bible_studies = reports.aggregate(
        average=Avg('bible_studies'))['average']
    avg_video_showings = reports.aggregate(
        average=Avg('video_showings'))['average']

    # avg_hours = total_hours/count_report

    nav = "Publishers"
    return render(response, "main/all-publishers/all-publishers-report-stats.html", {"publisher": publisher, "month_year": month_year, "category": category, "reports": reports, "count_report": count_report,
                                                                                    "total_hours": total_hours, "total_placements": total_placements, "total_return_visits": total_return_visits,
                                                                                    "total_bible_studies": total_bible_studies, "total_video_showings": total_video_showings,
                                                                                    "avg_hours": avg_hours, "avg_placements": avg_placements, "avg_return_visits": avg_return_visits,
                                                                                    "avg_bible_studies": avg_bible_studies, "avg_video_showings": avg_video_showings,
                                                                                    "nav": nav})

@authentication_required
@admin_privilege
@check_month_year
def monthly_pending_report(response):
    publisher = response.user
   
    if not response.session.get('publishers_report_month_year_id', None):
        return HttpResponseRedirect("/publishers/choose-category")

    month_year = MonthYear.objects.get(
        id=response.session['publishers_report_month_year_id'])
    Category = response.session['publishers_report_category']

    if Category == "one":
        category = "All Publishers"
        reports = PublisherReport.objects.filter(month_year_id=month_year.id)
        publishers = PublisherProfile.objects.filter(
            status="Active").exclude(groupName="Non-Publisher")
    elif Category == "two":
        category = "Regular Pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, pioneer="1")
        publishers = PublisherProfile.objects.filter(
            status="Active", pioneer="1").exclude(groupName="Non-Publisher")
    elif Category == "three":
        category = "Auxiliary Pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, pioneer="2")
        publishers = PublisherProfile.objects.filter(
            status="Active", pioneer="2").exclude(groupName="Non-Publisher")
    elif Category == "four":
        category = "Elders"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, appointment="1")
        publishers = PublisherProfile.objects.filter(
            status="Active", appointment="1").exclude(groupName="Non-Publisher")
    elif Category == "five":
        category = "Ministerial Servants"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, appointment="2")
        publishers = PublisherProfile.objects.filter(
            status="Active", appointment="2").exclude(groupName="Non-Publisher")

    count_publishers = publishers.count()
    count_report = reports.count()
    
    if reports.count() == 0:
        return HttpResponseRedirect("/publishers/choose-category")

    count_pending_reports = count_publishers - count_report

    if count_pending_reports > 1:
        pending_reports_tag = 'Pending reports'
    elif count_pending_reports == 1:
        pending_reports_tag = 'Pending report'
    else:
        pending_reports_tag = 'No Pending report'
        count_pending_reports = ''

    nav = "Publishers"
    return render(response, "main/all-publishers/pending-reports.html", {"publisher": publisher, "month_year": month_year, "category": category, "reports": reports, "publishers": publishers, "count_pending_reports": count_pending_reports, "pending_reports_tag": pending_reports_tag, "nav": nav})

@authentication_required
@admin_privilege
@check_month_year
def export_import_monthly_report(response):
    publisher = response.user
    
    if not response.session.get('publishers_report_month_year_id', None):
        return HttpResponseRedirect("/publishers/choose-category")

    month_year = MonthYear.objects.get(
        id=response.session['publishers_report_month_year_id'])
    Category = response.session['publishers_report_category']

    if Category == "one":
        category = "All Publishers"
        reports = PublisherReport.objects.filter(month_year_id=month_year.id)
        publishers = PublisherProfile.objects.filter(
            status="Active").exclude(groupName="Non-Publisher")
    elif Category == "two":
        category = "Regular Pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, pioneer="1")
        publishers = PublisherProfile.objects.filter(
            status="Active", pioneer="1").exclude(groupName="Non-Publisher")
    elif Category == "three":
        category = "Auxiliary Pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, pioneer="2")
        publishers = PublisherProfile.objects.filter(
            status="Active", pioneer="2").exclude(groupName="Non-Publisher")
    elif Category == "four":
        category = "Elders"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, appointment="1")
        publishers = PublisherProfile.objects.filter(
            status="Active", appointment="1").exclude(groupName="Non-Publisher")
    elif Category == "five":
        category = "Ministerial Servants"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, appointment="2")
        publishers = PublisherProfile.objects.filter(
            status="Active", appointment="2").exclude(groupName="Non-Publisher")

    count_publishers = publishers.count()
    count_report = reports.count()
    
    if reports.count() == 0:
        return HttpResponseRedirect("/publishers/choose-category")

    count_pending_reports = count_publishers - count_report

    if count_pending_reports > 1:
        pending_reports_tag = 'Pending reports'
    elif count_pending_reports == 1:
        pending_reports_tag = 'Pending report'
    else:
        pending_reports_tag = 'No Pending report'
        count_pending_reports = ''

    tests = PublisherReport.objects.filter(
        month_year_id="2")

    nav = "Publishers"
    return render(response, "main/all-publishers/export-import-reports.html", {"tests": tests, "publisher": publisher, "month_year": month_year, "category": category, "category_input": Category, "reports": reports, "publishers": publishers, "count_pending_reports": count_pending_reports, "pending_reports_tag": pending_reports_tag, "nav": nav})


def export_monthly_report_csv(request):
    if not request.session.get('publishers_report_month_year_id', None):
        return HttpResponseRedirect("/publishers/choose-category")

    month_year = MonthYear.objects.get(
        id=request.session['publishers_report_month_year_id'])
    Category = request.session['publishers_report_category']

    if Category == "one":
        category_tag = "all_publishers"
        reports = PublisherReport.objects.filter(month_year_id=month_year.id).values_list(
            'publisher_username', 'month', 'year', 'pioneer', 'appointment', 'placements', 'video_showings', 'hours', 'return_visits', 'bible_studies', 'comments')
    elif Category == "two":
        category_tag = "regular_pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, pioneer="1").values_list('publisher_username', 'month', 'year', 'pioneer', 'appointment', 'placements', 'video_showings', 'hours', 'return_visits', 'bible_studies', 'comments')
    elif Category == "three":
        category_tag = "auxiliary_pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, pioneer="2").values_list('publisher_username', 'month', 'year', 'pioneer', 'appointment', 'placements', 'video_showings', 'hours', 'return_visits', 'bible_studies', 'comments')
    elif Category == "four":
        category_tag = "elders"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, appointment="1").values_list('publisher_username', 'month', 'year', 'pioneer', 'appointment', 'placements', 'video_showings', 'hours', 'return_visits', 'bible_studies', 'comments')
    elif Category == "five":
        category_tag = "ministerial_servants"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, appointment="2").values_list('publisher_username', 'month', 'year', 'pioneer', 'appointment', 'placements', 'video_showings', 'hours', 'return_visits', 'bible_studies', 'comments')

    filename = category_tag+'_report_for_'+month_year.month+'_'+month_year.year

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="'+filename+'.csv"'

    writer = csv.writer(response)
    writer.writerow(['First Name', 'Middle Name', 'Last Name', 'Username', 'Month', 'Year', 'Pioneer', 'Appointment', 'Placements',
                    'Video Showings', 'Hours', 'Return Visits', 'Bible Studies', 'Comments'])

    for row in reports:
        publisher = PublisherProfile.objects.get(
                username=row[0])

        if row[3] == "1":
            pioneer = "R"
        elif row[4] == "2":
            pioneer = "A"    
        else:
            pioneer = ""

        if row[4] == "1":
            appointment = "E"
        elif row[4] == "2":
            appointment = "MS"
        else:
            appointment = ""

        writer.writerow([publisher.first_name, publisher.middle_name, publisher.last_name, row[0],
                        row[1], row[2], pioneer, appointment, row[5], row[6], row[7], row[8], row[9], row[10]])

    return response


def export_monthly_report_xls(request):

    if not request.session.get('publishers_report_month_year_id', None):
        return HttpResponseRedirect("/publishers/choose-category")

    month_year = MonthYear.objects.get(
        id=request.session['publishers_report_month_year_id'])
    Category = request.session['publishers_report_category']

    if Category == "one":
        category_tag = "all_publishers"
        reports = PublisherReport.objects.filter(month_year_id=month_year.id).values_list(
            'publisher_username', 'month', 'year', 'pioneer', 'appointment', 'placements', 'video_showings', 'hours', 'return_visits', 'bible_studies', 'comments')
    elif Category == "two":
        category_tag = "regular_pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, pioneer="1").values_list('publisher_username', 'month', 'year', 'pioneer', 'appointment', 'placements', 'video_showings', 'hours', 'return_visits', 'bible_studies', 'comments')
    elif Category == "three":
        category_tag = "auxiliary_pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, pioneer="2").values_list('publisher_username', 'month', 'year', 'pioneer', 'appointment', 'placements', 'video_showings', 'hours', 'return_visits', 'bible_studies', 'comments')
    elif Category == "four":
        category_tag = "elders"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, appointment="1").values_list('publisher_username', 'month', 'year', 'pioneer', 'appointment', 'placements', 'video_showings', 'hours', 'return_visits', 'bible_studies', 'comments')
    elif Category == "five":
        category_tag = "ministerial_servants"
        reports = PublisherReport.objects.filter(
            month_year_id=month_year.id, appointment="2").values_list('publisher_username', 'month', 'year', 'pioneer', 'appointment', 'placements', 'video_showings', 'hours', 'return_visits', 'bible_studies', 'comments')

    filename = category_tag+'_report_for_'+month_year.month+'_'+month_year.year

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="'+filename+'.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Reports')

    # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['FirstName', 'Middle Name', 'Last Name', 'Username', 'Month', 'Year', 'Pioneer', 'Appointment', 'Placements',
            'Video Showings', 'Hours', 'Return Visits', 'Bible Studies', 'Comments']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    # rows = User.objects.all().values_list('username', 'first_name', 'last_name', 'email')
    for row in reports:
        # row_num = 1
        row_num += 1

        #  convert tuple to list to update column
        my_list = list(row)
        if my_list[3] == "1":
            my_list[3] = "R"
        elif my_list[3] == "2":
            my_list[3] = "A"    
        else:
            my_list[3] = ""

        if my_list[4] == "1":
            my_list[4] = "E"
        elif my_list[4] == "2":
            my_list[4] = "MS"
        else:
            my_list[4] = ""

        row = tuple(my_list)
        publisher = PublisherProfile.objects.get(
                username=row[0])

        ws.write(row_num, 0, publisher.first_name, font_style)
        ws.write(row_num, 1, publisher.middle_name, font_style)
        ws.write(row_num, 2, publisher.last_name, font_style)
        ws.write(row_num, 3, row[0], font_style)
        ws.write(row_num, 4, row[1], font_style)
        ws.write(row_num, 5, row[2], font_style)
        ws.write(row_num, 6, row[3], font_style)
        ws.write(row_num, 7, row[4], font_style)
        ws.write(row_num, 8, row[5], font_style)
        ws.write(row_num, 9, row[6], font_style)
        ws.write(row_num, 10, row[7], font_style)
        ws.write(row_num, 11, row[8], font_style)
        ws.write(row_num, 12, row[9], font_style)
        ws.write(row_num, 13, row[10], font_style)

    wb.save(response)
    return response


def import_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES["excel_file"]
        
        # Check if the file is an Excel file (XLSX)
        if excel_file.name.endswith('.xlsx'):
            try:
                wb = load_workbook(excel_file)
                sheet = wb.active

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    
                    month_year = MonthYear.objects.get(
                    month=row[4], year=row[5])
                    
                    report = PublisherReport.objects.filter(
                    month_year_id=month_year.id, publisher_username=row[3])
                    
                    publisher_exist = PublisherProfile.objects.filter(username=row[3]).exists()
                    
                    if row[6] == "R":
                        pioneer = 1
                    elif row[6] == "A":  
                        pioneer = 2
                    else:
                        pioneer = ""
                        
                    if row[7] == "E":
                        appointment = 1
                    elif row[7] == "MS":  
                        appointment = 2
                    else:
                        appointment = ""         
                    
                    if publisher_exist:
                        if report.count() == 0:
                        
                            # Your data processing logic here
                            PublisherReport.objects.create(
                                publisher_username=row[3],
                                month=row[4],
                                year=row[5],
                                month_year_id=month_year.id,
                                pioneer=pioneer,
                                appointment=appointment,
                                placements="2",
                                video_showings="2",
                                hours="2",
                                return_visits="2",
                                bible_studies=row[12],
                                comments=row[13],
                                created_at=datetime.datetime.now()
                            )

                messages.success(request, 'XLSX file uploaded successfully!')
            except Exception as e:
                messages.error(request, f'Error processing Excel file: {str(e)}')

        # Check if the file is an Excel file (XLS)
        elif excel_file.name.endswith('.xls'):
            try:
                # Read the XLS file using xlrd
                book = xlrd.open_workbook(file_contents=excel_file.read())
                sheet = book.sheet_by_index(0)

                for row_num in range(1, sheet.nrows):
                    row = sheet.row_values(row_num)
                    
                    month_year = MonthYear.objects.get(
                    month=row[4], year=row[5])
                    
                    report = PublisherReport.objects.filter(
                    month_year_id=month_year.id, publisher_username=row[3])
                    
                    publisher_exist = PublisherProfile.objects.filter(username=row[3]).exists()
                    
                    if row[6] == "R":
                        pioneer = 1
                    elif row[6] == "A":  
                        pioneer = 2
                    else:
                        pioneer = ""
                        
                    if row[7] == "E":
                        appointment = 1
                    elif row[7] == "MS":  
                        appointment = 2
                    else:
                        appointment = "" 
                        
                    if publisher_exist:
                        if report.count() == 0:
                    
                            # Your data processing logic here
                            PublisherReport.objects.create(
                                publisher_username=row[3],
                                month=row[4],
                                year=row[5],
                                month_year_id=month_year.id,
                                pioneer=pioneer,
                                appointment=appointment,
                                placements="2",
                                video_showings="2",
                                hours="2",
                                return_visits="2",
                                bible_studies=row[12],
                                comments=row[13],
                                created_at=datetime.datetime.now()
                            )

                messages.success(request, 'XLS file uploaded successfully!')
            except Exception as e:
                messages.error(request, f'Error processing XLS file: {str(e)}')

        # Check if the file is a CSV file
        elif excel_file.name.endswith('.csv'):
            try:
                # Read the CSV file
                decoded_file = TextIOWrapper(excel_file.file, encoding='utf-8')
                reader = csv.reader(decoded_file)
                # Skip the header row
                next(reader, None)

                for row in reader:
                    
                    month_year = MonthYear.objects.get(
                    month=row[4], year=row[5])
                    
                    report = PublisherReport.objects.filter(
                    month_year_id=month_year.id, publisher_username=row[3])
                    
                    publisher_exist = PublisherProfile.objects.filter(username=row[3]).exists()
                    
                    if row[6] == "R":
                        pioneer = 1
                    elif row[6] == "A":  
                        pioneer = 2
                    else:
                        pioneer = ""
                        
                    if row[7] == "E":
                        appointment = 1
                    elif row[7] == "MS":  
                        appointment = 2
                    else:
                        appointment = "" 
                        
                    if publisher_exist:
                        if report.count() == 0:
                        
                            # Your data processing logic here
                            PublisherReport.objects.create(
                                publisher_username=row[3],
                                month=row[4],
                                year=row[5],
                                month_year_id=month_year.id,
                                pioneer=pioneer,
                                appointment=appointment,
                                placements="2",
                                video_showings="2",
                                hours="2",
                                return_visits="2",
                                bible_studies=row[12],
                                comments=row[13],
                                created_at=datetime.datetime.now()
                            )

                messages.success(request, 'CSV file uploaded successfully!')
            except Exception as e:
                messages.error(request, f'Error processing CSV file: {str(e)}')

        else:
            messages.error(request, 'Invalid file format. Please upload an Excel (XLSX) or CSV file.')

    else:
        messages.error(request, 'Invalid request. Please upload a file.')

    return redirect('/publishers/export-import-monthly-reports')

@authentication_required
@admin_privilege   
def delete_reports(request):
    publisher=request.user

    if not request.session.get('publishers_report_month_year_id', None):
        return HttpResponseRedirect("/publishers/choose-category")


    if request.method == 'POST':
        category = request.POST.get('category')
        if category == "one":
                # category = "All Publishers"
            reports = PublisherReport.objects.filter(month_year_id=request.POST.get('monthYearId'))
        elif category == "two":
            # category = "Regular Pioneers"
            reports = PublisherReport.objects.filter(
                month_year_id=request.POST.get('monthYearId'), pioneer="1")
        elif category == "three":
            # category = "Auxiliary Pioneers"
            reports = PublisherReport.objects.filter(
                month_year_id=request.POST.get('monthYearId'), pioneer="2")
        elif category == "four":
            # category = "Elders"
            reports = PublisherReport.objects.filter(
                month_year_id=request.POST.get('monthYearId'), appointment="1")
        elif category == "five":
            # category = "Ministerial Servants"
            reports = PublisherReport.objects.filter(
                month_year_id=request.POST.get('monthYearId'), appointment="2")
            
        reports.delete()
        return HttpResponse(1)
    else:
        return redirect("/logout")

def export_report__template_excel(request):

    if request.method == 'POST':
        month = request.POST.get('Month')
        year = request.POST.get('Year')
        Category = request.POST.get('Category')
        export_to = request.POST.get('export_to')
       
        
        if Category == "one":
            category_tag = "all_publishers"
            publishers = PublisherProfile.objects.filter(
                status="Active").exclude(groupName="Non-Publisher")
        elif Category == "two":
            category_tag = "regular_pioneers"
            publishers = PublisherProfile.objects.filter(
                status="Active", pioneer="1").exclude(groupName="Non-Publisher")
        elif Category == "three":
            category_tag = "auxiliary_pioneers"
            publishers = PublisherProfile.objects.filter(
                status="Active", pioneer="2").exclude(groupName="Non-Publisher")
        elif Category == "four":
            category_tag = "elders"
            publishers = PublisherProfile.objects.filter(
                status="Active", appointment="1").exclude(groupName="Non-Publisher")
        elif Category == "five":
            category_tag = "ministerial_servants"
            publishers = PublisherProfile.objects.filter(
                status="Active", appointment="2").exclude(groupName="Non-Publisher")
        filename = category_tag+'_report_template_for_'+month+'_'+year

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="'+filename+'.'+export_to+'"'
        
        if export_to == 'csv':
            writer = csv.writer(response)
            writer.writerow(['First Name', 'Middle Name', 'Last Name', 'Username', 'Month', 'Year', 'Pioneer', 'Appointment', 'Placements',
                            'Video Showings', 'Hours', 'Return Visits', 'Bible Studies', 'Comments'])

            for publisher in publishers:
        
                if publisher.pioneer == "1":
                    pioneer = "R"
                elif publisher.pioneer == "2":
                    pioneer = "A"    
                else:
                    pioneer = ""

                if publisher.appointment == "1":
                    appointment = "E"
                elif publisher.appointment == "2":
                    appointment = "MS"
                else:
                    appointment = ""

                writer.writerow([publisher.first_name, publisher.middle_name, publisher.last_name, publisher.username,
                                month, year, pioneer, appointment, "", "", "", "", "", ""])
        elif export_to == 'xls':
            wb = xlwt.Workbook(encoding='utf-8')
            ws = wb.add_sheet('Reports')

            # Sheet header, first row
            row_num = 0

            font_style = xlwt.XFStyle()
            font_style.font.bold = True

            columns = ['FirstName', 'Middle Name', 'Last Name', 'Username', 'Month', 'Year', 'Pioneer', 'Appointment', 'Placements',
                    'Video Showings', 'Hours', 'Return Visits', 'Bible Studies', 'Comments']

            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num], font_style)

            # Sheet body, remaining rows
            font_style = xlwt.XFStyle()

            # rows = User.objects.all().values_list('username', 'first_name', 'last_name', 'email')
            for publisher in publishers:
                # row_num = 1
                row_num += 1

                if publisher.pioneer == "1":
                    pioneer = "R"
                elif publisher.pioneer == "2":
                    pioneer = "A"    
                else:
                    pioneer = ""

                if publisher.appointment == "1":
                    appointment = "E"
                elif publisher.appointment == "2":
                    appointment = "MS"
                else:
                    appointment = ""
                    
                ws.write(row_num, 0, publisher.first_name, font_style)
                ws.write(row_num, 1, publisher.middle_name, font_style)
                ws.write(row_num, 2, publisher.last_name, font_style)
                ws.write(row_num, 3, publisher.username, font_style)
                ws.write(row_num, 4, month, font_style)
                ws.write(row_num, 5, year, font_style)
                ws.write(row_num, 6, pioneer, font_style)
                ws.write(row_num, 7, appointment, font_style)
                ws.write(row_num, 8, "", font_style)
                ws.write(row_num, 9, "", font_style)
                ws.write(row_num, 10, "", font_style)
                ws.write(row_num, 11, "", font_style)
                ws.write(row_num, 12, "", font_style)
                ws.write(row_num, 13, "", font_style)
            wb.save(response)    
        else: 
 
            # Create a new Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active

            # Write headers
            headers = ['FirstName', 'Middle Name', 'Last Name', 'Username', 'Month', 'Year', 'Pioneer', 'Appointment',
                    'Placements', 'Video Showings', 'Hours', 'Return Visits', 'Bible Studies', 'Comments']
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)
                

            # Write data
            for row_num, publisher in enumerate(publishers, 2):
                
                if publisher.pioneer == "1":
                    pioneer = "R"
                elif publisher.pioneer == "2":
                    pioneer = "A"    
                else:
                    pioneer = ""

                if publisher.appointment == "1":
                    appointment = "E"
                elif publisher.appointment == "2":
                    appointment = "MS"
                else:
                    appointment = ""
                    
                ws.cell(row=row_num, column=1, value=publisher.first_name)
                ws.cell(row=row_num, column=2, value=publisher.middle_name)
                ws.cell(row=row_num, column=3, value=publisher.last_name)
                ws.cell(row=row_num, column=4, value=publisher.username)
                ws.cell(row=row_num, column=5, value=month)
                ws.cell(row=row_num, column=6, value=year)
                ws.cell(row=row_num, column=7, value=pioneer)
                ws.cell(row=row_num, column=8, value=appointment)
                ws.cell(row=row_num, column=9, value="")
                ws.cell(row=row_num, column=10, value="")
                ws.cell(row=row_num, column=11, value="")
                ws.cell(row=row_num, column=12, value="")
                ws.cell(row=row_num, column=13, value="")
                ws.cell(row=row_num, column=14, value="")
             
            filename = category_tag+'_report_template_for_'+month+'_'+year
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="'+filename+'".xlsx'
            wb.save(response)       
        return response
    else:
        return redirect("/logout")
    
@authentication_required
@admin_privilege
@check_month_year   
def select_multiple_months(request):
    month_year_list = MonthYear.objects.filter().order_by('-id') 
    
    if request.method == 'POST':
        request.session['form_data'] = {
            'selected_values': request.POST.getlist('month_year_id[]'),
            'category':  request.POST.get('Category'),
        }
        return  HttpResponseRedirect("/publishers/multiple-months-report")
    else:
        if 'form_data' in request.session:
            del request.session['form_data']
            
    return render(request, "main/all-publishers/select-month-year.html", {"month_year_list": month_year_list})

@authentication_required
@admin_privilege
@check_month_year   
def multiple_months_report(request):
        form_data = request.session.get('form_data', {})
        selected_values = form_data.get('selected_values', [])
        category = form_data.get('category', '')
       
        if category == "one":
            category = "All Publishers"
            reports = PublisherReport.objects.filter(month_year_id__in=selected_values).values('publisher_username').annotate(
                count=Count('publisher_username'), total_hours=Sum('hours'), total_placements=Sum('placements'), total_return_visits=Sum('return_visits'), total_bible_studies=Sum('bible_studies'), total_video_showings=Sum('video_showings'))
        elif category == "two":
            category = "Regular Pioneers"
            reports = PublisherReport.objects.filter(
                month_year_id__in=selected_values, pioneer="1").values('publisher_username').annotate(
                    count=Count('publisher_username'), total_hours=Sum('hours'), total_placements=Sum('placements'), total_return_visits=Sum('return_visits'), total_bible_studies=Sum('bible_studies'), total_video_showings=Sum('video_showings'))
        elif category == "three":
            category = "Auxiliary Pioneers"
            reports = PublisherReport.objects.filter(
                month_year_id__in=selected_values, pioneer="2").values('publisher_username').annotate(
                    count=Count('publisher_username'), total_hours=Sum('hours'), total_placements=Sum('placements'), total_return_visits=Sum('return_visits'), total_bible_studies=Sum('bible_studies'), total_video_showings=Sum('video_showings'))
        elif category == "four":
            category = "Elders"
            reports = PublisherReport.objects.filter(
                month_year_id__in=selected_values, appointment="1").values('publisher_username').annotate(
                    count=Count('publisher_username'), total_hours=Sum('hours'), total_placements=Sum('placements'), total_return_visits=Sum('return_visits'), total_bible_studies=Sum('bible_studies'), total_video_showings=Sum('video_showings'))
        elif category == "five":
            category = "Ministerial Servants"
            reports = PublisherReport.objects.filter(
                month_year_id__in=selected_values, appointment="2").values('publisher_username').annotate(
                    count=Count('publisher_username'), total_hours=Sum('hours'), total_placements=Sum('placements'), total_return_visits=Sum('return_visits'), total_bible_studies=Sum('bible_studies'), total_video_showings=Sum('video_showings'))
  
        selected_month_years = MonthYear.objects.filter(id__in=selected_values)
        selected_month_number = selected_month_years.count()
        
        if selected_month_number == 0:
            messages.error(request, "ERROR: No month was checked")
            return  HttpResponseRedirect("/publishers/select-multiple-months-report")
        if reports.count() == 0:
            messages.error(request, "ERROR: No record was found under the selected month(s)")
            return  HttpResponseRedirect("/publishers/select-multiple-months-report")
        
        first_element = MonthYear.objects.get(id=selected_values[0])
        last_element = MonthYear.objects.get(id=selected_values[-1])

        context = {
            "selected_month_years": selected_month_years, 
            "category": category, 
            "reports": reports, 
            "selected_month_number": selected_month_number,
            "first_element": first_element,
            "last_element": last_element,
        }
        return render(request, "main/all-publishers/multiple-months-reports.html", context)

@authentication_required
@admin_privilege
@check_month_year   
def multiple_months_report_stats(request):
    form_data = request.session.get('form_data', {})
    selected_values = form_data.get('selected_values', [])
    Category = form_data.get('category', '')
    
    if Category == "one":
        category = "All Publishers"
        reports = PublisherReport.objects.filter(month_year_id__in=selected_values,)

    elif Category == "two":
        category = "Regular Pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id__in=selected_values, pioneer="1")

    elif Category == "three":
        category = "Auxiliary Pioneers"
        reports = PublisherReport.objects.filter(
            month_year_id__in=selected_values, pioneer="2")

    elif Category == "four":
        category = "Elders"
        reports = PublisherReport.objects.filter(
            month_year_id__in=selected_values, appointment="1")

    elif Category == "five":
        category = "Ministerial Servants"
        reports = PublisherReport.objects.filter(
            month_year_id__in=selected_values, appointment="2")
        

    number_of_publishers_reported = reports.values('publisher_username').annotate(
                    count=Count('publisher_username')).count()
    count_report = reports.count()
    total_hours = reports.aggregate(
        total=Sum('hours'))['total']
    total_placements = reports.aggregate(
        total=Sum('placements'))['total']
    total_return_visits = reports.aggregate(
        total=Sum('return_visits'))['total']
    total_bible_studies = reports.aggregate(
        total=Sum('bible_studies'))['total']
    total_video_showings = reports.aggregate(
        total=Sum('video_showings'))['total']

    selected_month_years = MonthYear.objects.filter(id__in=selected_values)
    selected_month_number = selected_month_years.count()
    
    if selected_month_number == 0 or count_report == 0:
        return  HttpResponseRedirect("/publishers/select-multiple-months-report")
    
    first_element = MonthYear.objects.get(id=selected_values[0])
    last_element = MonthYear.objects.get(id=selected_values[-1])
    
    if category == "All Publishers":
        tag = "Publishers"
    else:
        tag = category    

    context = {
        "selected_month_years": selected_month_years, 
        "category": category, 
        "reports": reports, 
        "selected_month_number": selected_month_number,
        "first_element": first_element,
        "last_element": last_element,
        "total_hours": total_hours,
        "total_placements": total_placements,
        "total_return_visits": total_return_visits,
        "total_bible_studies": total_bible_studies,
        "total_video_showings": total_video_showings,
        "number_of_publishers_reported": number_of_publishers_reported,
        "tag": tag,
    }
            
    return render(request, "main/all-publishers/multiple-months-reports-stats.html", context)    